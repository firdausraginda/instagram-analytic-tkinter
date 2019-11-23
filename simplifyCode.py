# -------------import-------------
import json
import operator
from operator import itemgetter
import datetime
import os
from collections import Counter 
from dateutil import tz
import calendar
import requests

# write excel
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill, GradientFill
wb = openpyxl.Workbook() 
sheet = wb.active 

# from openpyxl import Workbook
# from openpyxl.styles import Font
# book = Workbook()

def automateIgScraper(userName, destinationFile ,qtyPost):
    # destinationFile = "../data-ig/%s" % (userName)
    os.system('instagram-scraper --comments -m %s %s -d %s --retry-forever' % (qtyPost, userName, destinationFile))
    # destJsonFile = destinationFile + '/%s.json' % (userName)

def openFile(dataAkun):
    exists = os.path.isfile(dataAkun)
    if exists:
        with open(dataAkun, encoding="utf8") as f:
            d = json.load(f)
            data = d
    else:
        data = 'EMPTY DATA'

    return data

def findDay(date): 
    date = datetime.datetime.fromtimestamp(date).weekday()
    return calendar.day_name[date]

def findMonth(date):
    return datetime.datetime.fromtimestamp(date).month

def findHour(time):
    return datetime.datetime.fromtimestamp(time).hour

def convertDate(timeData):
    return datetime.datetime.fromtimestamp(timeData)

def removeEmoticon(kalimat):
    return kalimat.encode('ascii', 'ignore').decode('ascii')

def countLikes(data):
    counter = 0
    for i in range(0, len(data)):
        if isinstance(data[i]['count like'], str) == False:
            counter += data[i]['count like']
    
    return counter

def countComments(data):
    counter = 0
    for i in range(0, len(data)):
        if isinstance(data[i]['count comment'], str) == False:
            counter += data[i]['count comment']
    
    return counter

def countUsersThatComment(data):
    countUsersThatCommen = {}    
    for i in range(0, len(data)):
        for j in range(0, len(data[i]['Users That Comment'])):
            if data[i]['Users That Comment'][j] in countUsersThatCommen:
                countUsersThatCommen[data[i]['Users That Comment'][j]] += 1
            else:
                countUsersThatCommen[data[i]['Users That Comment'][j]] = 1

    k = Counter(countUsersThatCommen)
    highest = k.most_common(5)
    
    # for i in highest:
    #     print(i[0]," :",i[1]," ") 

    return highest

def getAllPostData(accOwner):
    allDataPost = []

    for i in range(0, len(accOwner['GraphImages'])):
        tempPost = {}
        idPost = accOwner['GraphImages'][i]['id'] 
        if(len(accOwner['GraphImages'][i]['edge_media_to_caption']['edges']) != 0):
            caption = removeEmoticon(accOwner['GraphImages'][i]['edge_media_to_caption']['edges'][0]['node']['text'])
            if(len(accOwner['GraphImages'][i]['tags']) != 0):
                separator = ', '
                tags = separator.join(accOwner['GraphImages'][i]['tags'])
            else:
                tags = ''
        else:
            caption = ''
            tags = ''
        img = accOwner['GraphImages'][i]['thumbnail_resources'][2]['src']
        hour = findHour(accOwner['GraphImages'][i]['taken_at_timestamp'])
        date = convertDate(accOwner['GraphImages'][i]['taken_at_timestamp'])
        day = findDay(accOwner['GraphImages'][i]['taken_at_timestamp'])
        month = findMonth(accOwner['GraphImages'][i]['taken_at_timestamp'])
        countLike = accOwner['GraphImages'][i]['edge_media_preview_like']['count']
        if (accOwner['GraphImages'][i]['comments_disabled'] == False):
            countComment = accOwner['GraphImages'][i]['edge_media_to_comment']['count']
            usersThatComment = []
            for j in range(0, len(accOwner['GraphImages'][i]['comments']['data'])):
                if accOwner['GraphImages'][i]['comments']['data'][j]['owner']['username'] != username:
                    usersThatComment.append(accOwner['GraphImages'][i]['comments']['data'][j]['owner']['username'])
        else:
            countComment = 0

        tempPost['id post'] = idPost
        tempPost['caption'] = caption
        tempPost['tags'] = tags
        tempPost['image'] = img
        tempPost['hour'] = hour
        tempPost['date'] = date
        tempPost['day'] = day
        tempPost['month'] = month
        tempPost['count like'] = countLike
        tempPost['count comment'] = countComment
        tempPost['Users That Comment'] = usersThatComment
        allDataPost.append(tempPost)

    return allDataPost

def getAllPostDataLocalServer(accOwner):
    allDataPost = []

    for i in range(0, len(accOwner)):
        tempPost = {}
        idPost = accOwner[i]['id'] 
        if(len(accOwner[i]['edge_media_to_caption']['edges']) != 0):
            caption = removeEmoticon(accOwner[i]['edge_media_to_caption']['edges'][0]['node']['text'])
            if(len(accOwner[i]['tags']) != 0):
                separator = ', '
                tags = separator.join(accOwner[i]['tags'])
            else:
                tags = ''
        else:
            caption = ''
            tags = ''
        img = accOwner[i]['thumbnail_resources'][2]['src']
        hour = findHour(accOwner[i]['taken_at_timestamp'])
        date = convertDate(accOwner[i]['taken_at_timestamp'])
        day = findDay(accOwner[i]['taken_at_timestamp'])
        month = findMonth(accOwner[i]['taken_at_timestamp'])
        countLike = accOwner[i]['edge_media_preview_like']['count']
        if (accOwner[i]['comments_disabled'] == False):
            countComment = accOwner[i]['edge_media_to_comment']['count']
            usersThatComment = []
            for j in range(0, len(accOwner[i]['comments']['data'])):
                if accOwner[i]['comments']['data'][j]['owner']['username'] != username:
                    usersThatComment.append(accOwner[i]['comments']['data'][j]['owner']['username'])
        else:
            countComment = 0

        tempPost['id post'] = idPost
        tempPost['caption'] = caption
        tempPost['tags'] = tags
        tempPost['image'] = img
        tempPost['hour'] = hour
        tempPost['date'] = date
        tempPost['day'] = day
        tempPost['month'] = month
        tempPost['count like'] = countLike
        tempPost['count comment'] = countComment
        tempPost['Users That Comment'] = usersThatComment
        allDataPost.append(tempPost)

    return allDataPost

def writeExcel(data, usersComments, totalPosts, totalLikes, totalComments):
    sheet.merge_cells('L4:M4')
    sheet.title = username
    sheet['A1'] = 'Id Post'
    sheet['B1'] = 'Date & Time'
    sheet['C1'] = 'Month'
    sheet['D1'] = 'Day'
    sheet['E1'] = 'Hour'
    sheet['F1'] = 'Interactions'
    sheet['G1'] = 'Likes'
    sheet['H1'] = 'Comments'
    sheet['I1'] = 'Tags'
    sheet['J1'] = 'Caption'
    sheet['K1'] = 'Image Url'

    sheet['M1'] = 'Total Posts'
    sheet['N1'] = 'Total Likes'
    sheet['O1'] = 'Total Comments'

    sheet['M4'] = 'Users With Highest Number of Comments'
    sheet['M5'] = 'Users'
    sheet['N5'] = 'Number of Comments'

    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet['D1'].font = Font(bold=True)
    sheet['E1'].font = Font(bold=True)
    sheet['F1'].font = Font(bold=True)
    sheet['G1'].font = Font(bold=True)
    sheet['H1'].font = Font(bold=True)
    sheet['I1'].font = Font(bold=True)
    sheet['J1'].font = Font(bold=True)
    sheet['K1'].font = Font(bold=True)

    sheet['M1'].font = Font(bold=True)
    sheet['N1'].font = Font(bold=True)
    sheet['O1'].font = Font(bold=True)

    sheet['M4'].font = Font(bold=True)
    sheet['M5'].font = Font(bold=True)
    sheet['N5'].font = Font(bold=True)
    sheet['M2'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['N2'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['O2'].alignment = Alignment(horizontal="center", vertical="center")
    
    for i in range(0, len(data)):
        sheet.cell(row=i+2, column=1).value = data[i]['id post']
        sheet.cell(row=i+2, column=2).value = data[i]['date']
        sheet.cell(row=i+2, column=3).value = data[i]['month']
        sheet.cell(row=i+2, column=4).value = data[i]['day']
        sheet.cell(row=i+2, column=5).value = data[i]['hour']
        sheet.cell(row=i+2, column=6).value = data[i]['count like'] + data[i]['count comment']
        sheet.cell(row=i+2, column=7).value = data[i]['count like']
        sheet.cell(row=i+2, column=8).value = data[i]['count comment']
        sheet.cell(row=i+2, column=9).value = data[i]['tags']
        sheet.cell(row=i+2, column=10).value = data[i]['caption']
        sheet.cell(row=i+2, column=11).value = data[i]['image']
    
    for i in range(0, len(usersComments)):
        sheet.cell(row=i+6, column=13).value = usersComments[i][0]
        sheet.cell(row=i+6, column=14).value = usersComments[i][1]

    sheet['M2'].value = totalPosts
    sheet['N2'].value = totalLikes
    sheet['O2'].value = totalComments

    wb.save('instagram-analytic-%s.xlsx' % (sheet.title))

def mainProg(username):
        # -----hosting locally-----
        url="http://localhost:3000/GraphImages"
        r = requests.get(url)
        data = r.json()

        # -----open file without hosting-----
        # data = openFile('../data-ig/%s/%s.json' % (username, username))

        if (data != 'EMPTY DATA'):
            # -----hosting locally-----
            allPostsData = getAllPostDataLocalServer(data)

            # -----open file without hosting-----
            # allPostsData = getAllPostData(data)

            totalPosts = len(allPostsData)
            totalLikes = countLikes(allPostsData)
            totalComments = countComments(allPostsData)
            countUsersComment = countUsersThatComment(allPostsData)
            writeExcel(allPostsData, countUsersComment, totalPosts, totalLikes, totalComments)
            # print(allPostsData)
        else:
            print('account is private or have not post anything yet')

# -----------------------------------------Pre-set-----------------------------------------
username = 'ra.ginda'
destFile = "../data-ig/%s" % (username)
maxData = 10
 
# -----------------------------------------Automate IG Scraper-----------------------------------------
# automateIgScraper(username, destFile, maxData)

# -----------------------------------------Main Program-----------------------------------------
mainProg(username)